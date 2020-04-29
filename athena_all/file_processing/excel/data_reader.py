from openpyxl import load_workbook
import glob

import pandas as pd
import numpy as np
import os 


class DataReader():

    def __init__(self, dataset_name):
        ''' Read from the given excel file. '''
        self.dataset_name = dataset_name
        self.workbook = load_workbook(filename=dataset_name, read_only=True, data_only=True)


    def get_all_sheets(self):
        ''' Load and save dataframes for every sheet in the excel file. '''
        dfs = []
        for sheet in self.workbook.sheetnames:
            ws = self.workbook[sheet]
            data = ws.values
            cols = next(data)[0:]
            data = list(data)
            df = pd.DataFrame(data, columns=cols)
            df.dropna(axis='columns',how='all', inplace=True)
            dfs.append(df)

        self.dfs = dfs
        return self.dfs

    def get_sheet_by_index(self, idx):
        ''' Only read and return a specific sheet from the excel file. '''
        if self.dfs: # If all sheets have been loaded just return from memory
            return self.dfs[idx]

        ws = self.workbook[self.workbook.sheetnames[idx]]
        data = ws.values
        cols = next(data)[0:]
        data = list(data)
        df = pd.DataFrame(data, columns=cols)
        df.dropna(axis='columns',how='all', inplace=True)
        return df